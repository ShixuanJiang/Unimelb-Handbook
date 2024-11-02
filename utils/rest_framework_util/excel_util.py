import openpyxl
from openpyxl import Workbook

__all__ = {
    'write_excel_file',
    'ExcelUtil'
}


def write_excel_file(row_data_list: list, file_name: str, title: str = None):
    wb = Workbook()
    sheet_name = None
    if title is not None:
        wb.create_sheet(title)
        sheet_name = title
    if sheet_name is None:
        sheet = wb[wb.sheetnames[0]]
        for row_data in row_data_list:
            sheet.append(row_data)
        wb.save(file_name)
    else:
        sheet = wb[sheet_name]
        for row_data in row_data_list:
            sheet.append(row_data)
        wb.save(file_name)


MAX_ROW_COUNT = 2000000  # 最大行数


class ExcelUtil(object):
    def __init__(self, excel_file_name, sheet_name: str = None):
        self.excel_name = excel_file_name
        try:
            self.workbook = openpyxl.load_workbook(self.excel_name)
            self.sheet_name = sheet_name
            if self.sheet_name is None:
                self.sheet = self.workbook[self.workbook.sheetnames[0]]
            else:
                self.sheet = self.workbook[sheet_name]
            self.row_count = self.sheet.max_row
            self.col_count = self.sheet.max_column
        except Exception as e:
            raise Exception(e)

    def read_data(self, row_count=None, col_count=None):
        title_list = []  # 标题列表
        response_data_info_list = []  # 数据列表
        if row_count is not None and row_count > MAX_ROW_COUNT:
            row_count = MAX_ROW_COUNT
        if row_count is not None:
            self.row_count = row_count
        if col_count is not None:
            self.col_count = col_count
        if self.row_count < 1 or self.col_count < 1:
            return title_list, response_data_info_list
        for index, data in enumerate(self.sheet):
            if index == self.row_count:
                break
            data_list = []
            for i in range(0, self.col_count):
                if index == 0:
                    title_list.append(data[i].value)
                else:
                    data_list.append(data[i].value)
            if index != 0:
                response_data_info_list.append(data_list)
        return title_list, response_data_info_list


if __name__ == '__main__':
    excel_file = ExcelUtil('123.xlsx')
    title_data, row_data = excel_file.read_data()
    write_excel_file(row_data, 'xssx.xlsx')
